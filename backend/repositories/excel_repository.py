import os
import openpyxl
from openpyxl import Workbook
from typing import List, Dict, Any, Optional
from schemas.patient_schema import PatientJSON

COLUMNS = [
    "Patient ID",
    "UUID",
    "Group",
    "Name",
    "Age Group",
    "Gender",
    "Phone",
    "Address",
    "Smoking",
    "Alcohol Consumption",
    "Chewing Habit",
    "Medical Condition",
    "Allergies",
    "Notes",
    "Date Created",
    "Last Updated",
    "Status",
    "Right Thumb Image Path",
    "Right Thumb Annotated Path",
    "Right Index Image Path",
    "Right Index Annotated Path",
    "Right Middle Image Path",
    "Right Middle Annotated Path",
    "Right Ring Image Path",
    "Right Ring Annotated Path",
    "Right Little Image Path",
    "Right Little Annotated Path",
    "Left Thumb Image Path",
    "Left Thumb Annotated Path",
    "Left Index Image Path",
    "Left Index Annotated Path",
    "Left Middle Image Path",
    "Left Middle Annotated Path",
    "Left Ring Image Path",
    "Left Ring Annotated Path",
    "Left Little Image Path",
    "Left Little Annotated Path"
]

class ExcelRepository:
    """Manages the patients.xlsx file using OpenPyXL for indexing and fast lookup."""

    def __init__(self, data_dir: str = None):
        if data_dir is None:
            data_dir = os.environ.get("FING_DATA_DIR")
            
        if data_dir is None:
            import sys
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            self.data_dir = os.path.join(base_dir, "data")
        else:
            self.data_dir = os.path.abspath(data_dir)
            
        self.excel_path = os.path.join(self.data_dir, "patients.xlsx")
        os.makedirs(self.data_dir, exist_ok=True)
        
        # Self-healing index: check if file has correct headers
        should_recreate = False
        if os.path.exists(self.excel_path):
            try:
                wb = openpyxl.load_workbook(self.excel_path, read_only=True)
                if "Patients" in wb.sheetnames:
                    ws = wb["Patients"]
                    row_1 = next(ws.iter_rows(max_row=1, values_only=True))
                    if list(row_1) != COLUMNS:
                        should_recreate = True
                else:
                    should_recreate = True
            except Exception:
                should_recreate = True
                
        self._init_excel_file(force_recreate=should_recreate)

    def _init_excel_file(self, force_recreate: bool = False) -> None:
        """Creates the Excel file with correct columns if it does not exist or is mismatching."""
        if force_recreate and os.path.exists(self.excel_path):
            try:
                os.remove(self.excel_path)
                print(f"[OK] Removed mismatching patients index Excel at {self.excel_path}")
            except Exception as e:
                print(f"[WARN] Failed to delete mismatching Excel file: {e}. Attempting in-place overwrite.")
                try:
                    wb = openpyxl.load_workbook(self.excel_path)
                    if "Patients" not in wb.sheetnames:
                        ws = wb.create_sheet("Patients")
                    else:
                        ws = wb["Patients"]
                    # Clear all rows
                    ws.delete_rows(1, ws.max_row + 1)
                    # Write headers
                    for col_idx, col_name in enumerate(COLUMNS, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    # Style
                    ws.row_dimensions[1].height = 20
                    for col in range(1, len(COLUMNS) + 1):
                        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
                    self._save_workbook_safe(wb)
                    print(f"[OK] Overwrote Excel headers in-place successfully.")
                    return
                except Exception as ex:
                    print(f"[ERROR] Failed to overwrite Excel headers in-place: {ex}")
                
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Patients"
            
            # Write headers
            for col_idx, col_name in enumerate(COLUMNS, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
                
            # Apply basic header styling
            ws.row_dimensions[1].height = 20
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
                
            self._save_workbook_safe(wb)
            print(f"[OK] Created new patients index Excel at {self.excel_path}")

    def _save_workbook_safe(self, wb) -> None:
        """Saves the workbook with retries to handle locked files gracefully."""
        import time
        for attempt in range(8):
            try:
                wb.save(self.excel_path)
                return
            except PermissionError as pe:
                if attempt == 7:
                    print(f"[ERROR] Failed to save Excel index due to lock: {pe}. Skipping Excel save fallback.")
                    return
                print(f"[WARN] Excel index locked, retrying in 0.3s...")
                time.sleep(0.3)

    def upsert_patient(self, patient: PatientJSON) -> None:
        """
        Inserts or updates a patient row in the Excel sheet.
        Matches by UUID (Column B / Column 2).
        """
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Patients"]
        
        # Search for UUID
        uuid_col_idx = 2
        found_row_idx = -1
        
        # Iterate over all rows starting from row 2
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=uuid_col_idx).value
            if cell_val == patient.uuid:
                found_row_idx = row
                break
                
        # Get fingerprint paths in order of COLUMNS
        fp_paths = []
        all_fingers = [
            "right_thumb", "right_index", "right_middle", "right_ring", "right_little",
            "left_thumb", "left_index", "left_middle", "left_ring", "left_little"
        ]
        for f in all_fingers:
            db_fp = patient.fingerprints.get(f)
            if db_fp:
                fp_paths.append(db_fp.image_path if db_fp.image_path else "NA")
                fp_paths.append(db_fp.annotated_image_path if db_fp.annotated_image_path else "NA")
            else:
                fp_paths.append("NA")
                fp_paths.append("NA")
                
        # Prep row values
        row_values = [
            patient.patient_id,
            patient.uuid,
            patient.group,
            patient.name,
            patient.age_group,
            patient.gender,
            patient.phone,
            patient.address,
            patient.additional_details.smoking,
            patient.additional_details.alcohol_consumption,
            patient.additional_details.chewing_habit,
            patient.additional_details.medical_condition,
            patient.additional_details.allergies,
            patient.additional_details.notes,
            patient.created_date,
            patient.updated_date,
            patient.status
        ] + fp_paths
        
        # Clean values: strip spaces and fallback empty values to "NA"
        row_values = [
            str(val).strip() if (val is not None and str(val).strip() != "") else "NA"
            for val in row_values
        ]
        
        if found_row_idx != -1:
            # Update existing row
            print(f"[INFO] Updating Excel row {found_row_idx} for patient {patient.patient_id}")
            for col_idx, val in enumerate(row_values, 1):
                ws.cell(row=found_row_idx, column=col_idx, value=val)
        else:
            # Append new row
            new_row_idx = ws.max_row + 1
            print(f"[INFO] Appending new Excel row {new_row_idx} for patient {patient.patient_id}")
            for col_idx, val in enumerate(row_values, 1):
                ws.cell(row=new_row_idx, column=col_idx, value=val)
                
        self._save_workbook_safe(wb)

    def search_patients(
        self, 
        name: Optional[str] = None, 
        phone: Optional[str] = None, 
        patient_id: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Searches the Excel index for patients.
        Performs case-insensitive substring matching.
        """
        if not os.path.exists(self.excel_path):
            return []
            
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        if "Patients" not in wb.sheetnames:
            return []
            
        ws = wb["Patients"]
        results = []
        
        # Read header row to map columns
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        
        # Match indexes
        try:
            id_idx = headers.index("Patient ID")
            uuid_idx = headers.index("UUID")
            group_idx = headers.index("Group")
            name_idx = headers.index("Name")
            phone_idx = headers.index("Phone")
            status_idx = headers.index("Status")
        except ValueError as e:
            print(f"[ERROR] Excel header corruption: {e}")
            return []
            
        # Iterate through rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(id_idx, uuid_idx, group_idx, name_idx, phone_idx, status_idx):
                continue
                
            row_id = str(row[id_idx] or "").strip()
            row_uuid = str(row[uuid_idx] or "").strip()
            row_group = str(row[group_idx] or "").strip()
            row_name = str(row[name_idx] or "").strip()
            row_phone = str(row[phone_idx] or "").strip()
            row_status = str(row[status_idx] or "").strip()
            
            # Check filters
            match = True
            
            if name and name.strip():
                if name.strip().lower() not in row_name.lower():
                    match = False
                    
            if phone and phone.strip():
                if phone.strip().lower() not in row_phone.lower():
                    match = False
                    
            if patient_id and patient_id.strip():
                if patient_id.strip().lower() not in row_id.lower():
                    match = False
            
            # If all provided filters match, append
            if match:
                results.append({
                    "uuid": row_uuid,
                    "patient_id": row_id,
                    "group": row_group,
                    "name": row_name,
                    "phone": row_phone,
                    "status": row_status
                })
                
        return results

    def clear_excel_rows(self) -> None:
        """Clears all data rows, keeping only headers."""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Patients"]
        # Delete rows starting from row 2
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        self._save_workbook_safe(wb)
